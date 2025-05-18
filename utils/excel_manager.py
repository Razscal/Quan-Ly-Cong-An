import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.person import Person
from models.award import Award

def create_excel_template(file_path, columns):
    """
    Create an Excel template with the specified columns.
    
    Args:
        file_path: Path to save the Excel file
        columns: List of column names
    """
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Nhiệm vụ"
    
    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Add header row
    for col_idx, column in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = column
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        
        # Set column width
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 20
    
    # Add 10 empty rows for data
    for row_idx in range(2, 12):
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
    
    # Save the workbook
    wb.save(file_path)

def merge_excel_files(input_files, output_file):
    """
    Merge multiple Excel files into one by appending all rows.
    
    Args:
        input_files: List of input Excel file paths
        output_file: Path to save the merged Excel file
    """
    # Check if input files exist
    for file in input_files:
        if not os.path.exists(file):
            raise FileNotFoundError(f"File not found: {file}")
    
    # Validate input
    if not input_files:
        raise ValueError("No input files provided")
    
    # Read all dataframes with explicit sheet name
    all_data = []
    headers = None
    
    # First, get headers from the first file to ensure consistency
    first_df = pd.read_excel(input_files[0])
    headers = list(first_df.columns)
    all_data.append(first_df)
    
    # Process remaining files
    for file in input_files[1:]:
        try:
            # Read with the same column structure
            df = pd.read_excel(file)
            
            # Check if columns match, if not, try to align them
            if list(df.columns) != headers:
                # Reorder columns if possible, or fill missing ones
                df = df.reindex(columns=headers, fill_value=None)
            
            all_data.append(df)
        except Exception as e:
            print(f"Error reading {file}: {str(e)}")
    
    # Concatenate all dataframes, preserving all rows
    if all_data:
        try:
            # Use axis=0 to append rows (default)
            merged_df = pd.concat(all_data, ignore_index=True, sort=False)
            
            # Apply styling to the output file
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                merged_df.to_excel(writer, index=False, sheet_name="Nhiệm vụ")
                
                # Apply styling to the header
                workbook = writer.book
                worksheet = writer.sheets["Nhiệm vụ"]
                
                # Define styles
                header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
                
                # Apply styles to header row
                for col_idx, column in enumerate(merged_df.columns, start=1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                    
                    # Set column width
                    column_letter = get_column_letter(col_idx)
                    worksheet.column_dimensions[column_letter].width = 20
                
                # Apply borders to all data cells
                for row_idx in range(2, len(merged_df) + 2):  # +2 because Excel is 1-indexed and we have a header row
                    for col_idx in range(1, len(merged_df.columns) + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.border = border
                        
        except Exception as e:
            raise Exception(f"Error merging files: {str(e)}")
    else:
        # Create an empty Excel file if no data
        wb = Workbook()
        ws = wb.active
        ws.title = "Nhiệm vụ"
        wb.save(output_file)

def import_excel_data(file_path, task, session):
    """
    Import data from an Excel file into the database.
    
    Args:
        file_path: Path to the Excel file
        task: Task object to associate with the imported data
        session: Database session
    """
    # Check if file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    
    # Read Excel file
    df = pd.read_excel(file_path)
    
    # Check if the dataframe is empty
    if df.empty:
        return
    
    # Get the name column (assuming the first column is always the name)
    name_column = df.columns[0]
    
    # Clear existing people for this task to avoid duplicates when reimporting
    # This is optional and depends on whether you want to replace or append data
    # Uncomment the following lines if you want to clear existing data before import
    # existing_people = session.query(Person).filter(Person.task_id == task.id).all()
    # for person in existing_people:
    #     session.delete(person)
    # session.flush()
    
    # Process each row
    for _, row in df.iterrows():
        name = row[name_column]
        if not pd.isna(name) and name.strip():
            # Check if person already exists
            person = session.query(Person).filter(
                Person.name == name.strip(),
                Person.task_id == task.id
            ).first()
            
            # Create a new person if not exists
            if not person:
                person = Person(name=name.strip(), task_id=task.id)
                session.add(person)
                session.flush()  # Flush to get the person ID
            
            # Process all other columns as potential award columns
            # Skip the first column (name) and process all others
            for col in df.columns[1:]:
                award_text = row[col]
                if not pd.isna(award_text) and str(award_text).strip():
                    # Try to extract year from award text (format: "Award Name (Year)")
                    award_name = str(award_text).strip()
                    award_year = datetime.now().year  # Default to current year
                    
                    # Check if award has year in parentheses
                    if "(" in award_name and ")" in award_name:
                        try:
                            year_text = award_name.split("(")[1].split(")")[0]
                            if year_text.isdigit():
                                award_year = int(year_text)
                                award_name = award_name.split("(")[0].strip()
                        except:
                            pass
                    
                    # Use column name as award category if needed
                    award_category = col
                    
                    # Check if award already exists
                    award = session.query(Award).filter(
                        Award.name == award_name,
                        Award.year == award_year,
                        Award.person_id == person.id
                    ).first()
                    
                    # Create a new award if not exists
                    if not award:
                        award = Award(
                            name=award_name,
                            year=award_year,
                            person_id=person.id
                        )
                        session.add(award)
